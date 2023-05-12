import openpyxl
def write_xls(file_path,lis):
    list_alis=[]
    list_blis=[]
    for i in range(0,len(lis)):
        if(i%2==0):
            list_blis.append(lis[i])
        if(i%2==1):
            list_alis.append(lis[i])
    #print(list_alis)
    #print(list_blis)

    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)
    # 获取“Sheet1”工作表
    ws = workbook['Sheet1']
    #最大行数：
    max_rol=ws.max_row

    # 在第1行第1列中添加值
    for i in range (0,len(list_alis)):
        rol_num=max_rol+i+1
        ws.cell(row=rol_num, column=2).value =list_alis[i]
        ws.cell(row=rol_num, column=1).value = list_blis[i]

    # 保存更改
    workbook.save('xlsx/optb.xlsx')

#write_xls(file_path,lis)